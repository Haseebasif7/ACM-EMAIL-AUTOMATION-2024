import openpyxl;
from letterFunction import getLetterWithAttributes;
path ="C:\\Users\\CoreCom\\Downloads\\Automate.xlsx"
obj = openpyxl.load_workbook(path)
sheet_obj = obj.active

for i in range(2,sheet_obj.max_row + 1):
  name = sheet_obj.cell(row=i,column=4)
  team = sheet_obj.cell(row=i,column=6)
  position = sheet_obj.cell(row=i,column=8)
  getLetterWithAttributes(name,position,team)
 
