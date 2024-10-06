import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl

gmail_user = ''
gmail_password = ''

workbook = openpyxl.load_workbook('C:/Users/CoreCom/Downloads/IDs.xlsx')
sheet = workbook.active

image_url = 'https://drive.google.com/uc?export=view&id=1XRpg-Sc-2YJLAaseX3J2yTD-2Vf5GAaA'

for i in range(2, sheet.max_row + 1): 
    recipient_email = sheet.cell(row=i, column=1).value  


    if recipient_email is not None:
        msg = MIMEMultipart('alternative')
        msg['From'] = gmail_user
        msg['To'] = recipient_email
        msg['Subject'] = 'Testing HTML Template Email'

        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
        <style>
            body {{
                font-family: Arial, sans-serif;
                margin: 0;
                padding: 0;
                background-color: #f4f4f4;
            }}
            .email-container {{
            width: 80%;
            max-width: 500px;
            margin: 0 auto;
            background-color: #ffffff;
            padding: 20px;
            border: 5px solid rgb(13, 95, 145);
            border-radius: 5px;
            }}
            .email-header {{
                padding: 10px;
                text-align: center;
                color: rgb(0, 0, 0);
            }}
            .email-body {{
                padding: 20px;
            }}
            .email-body h2 {{
                color: #333;
            }}
            .email-footer {{
                text-align: center;
                font-size: 12px;
                color: #999;
                padding: 20px 0;
                border-top: 1px solid #e0e0e0;
            }}
            .email-footer a {{
                text-decoration: none;
            }}
            
        </style> 
        </head>

        <body>
            <div class="email-container">
                <div class="email-header">
                    <img src="{image_url}" style="width: 10vw;" alt="ACM Logo">
                    <h1>Welcome to ACM</h1>
                </div>

                <!-- Body -->
                <div class="email-body">
                    <h2>Hello!</h2>
                    <p>Welcome to the ACM community. We are excited to have you!</p>
                </div>

                <!-- Footer -->
                <div class="email-footer">
                    <p>Regards,<br>ACM Team</p>
                </div>
            </div>
        </body>
</html>
        """

        msg.attach(MIMEText(html, 'html'))

        try:
            server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
            server.login(gmail_user, gmail_password)
            server.sendmail(gmail_user, recipient_email, msg.as_string())
            server.close()

            print(f"Email sent successfully to {recipient_email}!")
        except Exception as e:
            print(f"Failed to send email to {recipient_email}: {e}")
    else:
        print(f"No valid email found at row {i}. Skipping...")
